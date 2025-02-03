import openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.styles import Alignment, Border, Side, PatternFill
from datetime import datetime
import json
import sys
import yaml  # PyYAML を利用。事前に pip install pyyaml を実施してください。
import os  # ファイル名操作用
import re

class HeaderDefinition:
    """
    ヘッダー部の情報を保持するクラス。
    対象シート内の A列が "#JSON_START" となっている行からヘッダ情報を抽出し、
    各列ごとのキー階層情報（例： [{"key": "mailSet", "arr": None}, {"key": "mailTemplates", "arr": "$1"}, … ]）を保持します。
    """
    def __init__(self, ws):
        self.ws = ws
        self.hierarchy = {}  # key: 列文字（例 "B", "C", ...）、value: ヘッダー定義リスト
        self.parse_headers()

    def parse_headers(self):
        header_rows = []
        for row in self.ws.iter_rows(min_row=1, max_row=self.ws.max_row):
            cell_a = row[0].value
            if isinstance(cell_a, str) and cell_a.strip() == "#JSON_START":
                header_rows.append(row)
        
        if not header_rows:
            print("警告：ヘッダー行（#JSON_START）が見つかりませんでした。")
            return
        
        for header_row in header_rows:
            for cell in header_row:
                if cell.column == 1:
                    continue  # A列はヘッダー識別用としてスキップ
                cell_value = cell.value
                # 統合セルの場合は、先頭セルの値を取得
                if cell_value is None:
                    for merged_range in self.ws.merged_cells.ranges:
                        if cell.coordinate in merged_range:
                            cell_value = self.ws.cell(row=merged_range.min_row, column=merged_range.min_col).value
                            break
                if cell_value and isinstance(cell_value, str) and cell_value.strip().startswith("#"):
                    key_text = cell_value.strip().lstrip("#").strip()
                    # 正規表現で「キー名」「配列指示子（数字部分）」およびグループ名を抽出する
                    m = re.match(r"^(.*?)\$(\d+)(.*)$", key_text)
                    if m:
                        base_key = m.group(1).strip()
                        digit = m.group(2).strip()
                        group_name = m.group(3).strip()
                        # グループ名があればキー名に結合する例
                        if group_name:
                            key_name = base_key + group_name
                        else:
                            key_name = base_key
                        array_directive = f"${digit}"
                    else:
                        key_name = key_text
                        array_directive = None
                    col_letter = get_column_letter(cell.column)
                    if col_letter not in self.hierarchy:
                        self.hierarchy[col_letter] = []
                    self.hierarchy[col_letter].append({"key": key_name, "arr": array_directive})

class RowDataProcessor:
    """
    ヘッダ情報（HeaderDefinition）をもとに、複製シート内のデータ行（A列が "#JSON_DATA" の行）を
    ネスト構造の辞書へ変換するクラスです。
    
    メイン行では、各列のセルの値を対応するヘッダー定義に沿って配置し、
    継続行（B列などに "$" 指定がある行）の場合は、同じ親パス内でグループ化して配列要素として追加します。
    """
    def __init__(self, ws, header_definition):
        self.ws = ws
        self.header_mapping = header_definition.hierarchy
        self.rows = list(ws.iter_rows(min_row=1, max_row=ws.max_row))

    def _nested_update(self, d, keys, value):
        if not keys:
            return
        key = keys[0]
        if len(keys) == 1:
            d[key] = value
        else:
            if key not in d or not isinstance(d[key], dict):
                d[key] = {}
            self._nested_update(d[key], keys[1:], value)

    def _update_record_with_header(self, record, header_defs, value):
        """
        ヘッダー定義に沿って record の指定位置に value を設定する。
        もし配列指定子があり、かつ末端の場合は、その配列に直接値を追加します。
        """
        arr_index = None
        for idx, level in enumerate(header_defs):
            if level.get("arr"):
                arr_index = idx
                break

        if arr_index is None:
            # 配列指定子がない場合は単純なネスト更新
            current = record
            for j, level in enumerate(header_defs):
                key = level["key"]
                if j == len(header_defs) - 1:
                    current[key] = value
                else:
                    if key not in current or not isinstance(current[key], dict):
                        current[key] = {}
                    current = current[key]
        else:
            # 配列指定子が存在する場合
            parent_path = [l["key"] for l in header_defs[:arr_index]]
            array_key = header_defs[arr_index]["key"]
            sub_keys = [l["key"] for l in header_defs[arr_index+1:]]
            current = record
            for key in parent_path:
                if key not in current or not isinstance(current[key], dict):
                    current[key] = {}
                current = current[key]
            if not sub_keys:
                # 末端の場合は、直接値を配列に追加
                if array_key not in current or not isinstance(current[array_key], list):
                    current[array_key] = []
                current[array_key].append(value)
            else:
                if array_key not in current:
                    current[array_key] = []
                if len(current[array_key]) == 0:
                    current[array_key].append({})
                self._nested_update(current[array_key][0], sub_keys, value)

    def process_rows(self):
        data_list = []
        i = 0
        while i < len(self.rows):
            row = self.rows[i]
            marker = row[0].value
            if marker is None or not isinstance(marker, str) or not marker.strip().startswith("#JSON_DATA"):
                i += 1
                continue

            new_record = {}
            main_row = row
            main_row_num = main_row[0].row

            # メイン行：各列データをヘッダー定義に沿って更新
            for col_letter, header_defs in self.header_mapping.items():
                col_idx = column_index_from_string(col_letter)
                cell_val = self.ws.cell(row=main_row_num, column=col_idx).value
                self._update_record_with_header(new_record, header_defs, cell_val)
            i += 1

            # 継続行の処理（複数列にわたる配列指定子の更新）
            while i < len(self.rows):
                next_row = self.rows[i]
                directive = None
                if len(next_row) > 1:
                    directive = next_row[1].value
                if directive is None:
                    break

                cont_updates = {}  # key: (親キータプル, 配列キー)、value: 更新用（dictまたはlist）
                cont_row_num = next_row[0].row
                for col_letter, header_defs in self.header_mapping.items():
                    for idx, level in enumerate(header_defs):
                        if level.get("arr") == f'${directive}':
                            parent_path = tuple(l["key"] for l in header_defs[:idx])
                            array_key = header_defs[idx]["key"]
                            sub_keys = [l["key"] for l in header_defs[idx+1:]]
                            col_idx = column_index_from_string(col_letter)
                            cell_val = self.ws.cell(row=cont_row_num, column=col_idx).value
                            key = (parent_path, array_key)
                            if key not in cont_updates:
                                if sub_keys:
                                    cont_updates[key] = {}
                                else:
                                    cont_updates[key] = []
                            if sub_keys:
                                self._nested_update(cont_updates[key], sub_keys, cell_val)
                            else:
                                cont_updates[key].append(cell_val)
                            break

                # 継続行の更新内容を該当する親階層の配列に追加する
                for (parent_path, array_key), update_val in cont_updates.items():
                    current = new_record
                    for k in parent_path:
                        if k not in current or not isinstance(current[k], dict):
                            current[k] = {}
                        current = current[k]
                    if array_key not in current:
                        current[array_key] = []
                    if isinstance(update_val, list):
                        current[array_key].extend(update_val)
                    else:
                        current[array_key].append(update_val)
                i += 1

            data_list.append(new_record)
        return data_list

class ExcelJsonConverter:
    def __init__(self, input_file):
        """
        Excelファイルを読み込み、対象シートを決定して内部状態として保持します。
        対象シートの決定ルール：
          - シート名が "JSON_START" のシートがあればそれを使用
          - なければシートがひとつの場合そのシートを使用
        """
        self.input_file = input_file
        self.wb = openpyxl.load_workbook(input_file)
        self.target_ws = self._select_target_sheet()
        self.duplicate_ws = None  # 複製したシートを後で設定

    def _select_target_sheet(self):
        if "JSON_START" in self.wb.sheetnames:
            return self.wb["JSON_START"]
        elif len(self.wb.sheetnames) == 1:
            return self.wb.active
        else:
            print("エラー：複数シートありますが、'JSON_START' という名前のシートが見つかりません。")
            sys.exit(1)

    def duplicate_sheet(self):
        """
        対象シートを複製し、シート名を「JSON_YYYYMMDD_HHMMSS」に変更します。
        複製したシートは self.duplicate_ws に保持されます。
        """
        self.duplicate_ws = self.wb.copy_worksheet(self.target_ws)
        dt_sheet_name = datetime.now().strftime("%Y%m%d_%H%M%S")
        self.duplicate_ws.title = "JSON_" + dt_sheet_name

    def transform_data_markers(self):
        """
        複製シート内で、A列に記載されている
          #JSON_DATA_START ～ #JSON_DATA_END のブロック（もしくは
          #JSON_DATA_START からシート最終行まで）
        の各行について、セルの値を通常は "#JSON_DATA" に置換（補完）します。
        ※ セルの値が "#NOT" であれば、その行は変更しません。
        ただし、セルの値が "*" で始まる場合（例: "*$1"）は継続行として扱うため、そのまま保持します。
        ※ 本例では、継続行の判断はB列の値が「$」で始まるかおよび空セルになるまでで行います。
        """
        ws = self.duplicate_ws
        inside_block = False
        max_row = ws.max_row
        for i in range(1, max_row + 1):
            cell = ws.cell(row=i, column=1)
            marker = ""
            if cell.value and isinstance(cell.value, str):
                marker = cell.value.strip()
            if marker == "#JSON_DATA_START":
                inside_block = True
                cell.value = "#JSON_DATA"
            elif marker == "#JSON_DATA_END":
                cell.value = "#JSON_DATA"
                inside_block = False
            elif inside_block:
                if marker.startswith("*"):
                    continue
                elif marker != "#NOT":
                    cell.value = "#JSON_DATA"

    def convert_sheet_to_json(self):
        """
        複製シート内のデータ行（A列が "#JSON_DATA" の行）を、
        ヘッダー定義および行データ処理オブジェクトを用いて変換します。
        """
        header_def = HeaderDefinition(self.duplicate_ws)
        row_processor = RowDataProcessor(self.duplicate_ws, header_def)
        return row_processor.process_rows()

    def _set_title_cell(self, ws, row, column, text, fill, border):
        """
        指定されたセルにタイトル用の値とスタイルを設定して返します。
        """
        cell = ws.cell(row=row, column=column, value=text)
        cell.fill = fill
        cell.border = border
        cell.alignment = Alignment(horizontal="center", vertical="center")
        return cell

    def _setup_output_cell(self, ws, row, column, text, border):
        """
        指定されたセルにテキストを設定し、テキスト形式(@)、
        折り返し表示、罫線を設定したセルオブジェクトを返します。
        """
        cell = ws.cell(row=row, column=column, value=text)
        cell.number_format = "@"
        cell.alignment = Alignment(wrap_text=True)
        cell.border = border
        return cell

    def output_json_yaml(self, json_data, offset=10):
        """
        変換したJSONデータ（リスト）を JSON 形式と YAML 形式に変換し、
        複製シートの最終行から offset 行下に出力します。
        JSON は A列、YAML は B列に出力し、各セルの書式をテキスト形式かつ折り返し表示に設定します。
        また、JSON/YAMLの出力セルの１つ上のセルにタイトル（出力JSON、出力YAML）を
        背景を明るい紫色にし、タイトルと内容を罫線で囲むようにします。
        """
        ws = self.duplicate_ws
        json_text = json.dumps(json_data, ensure_ascii=False, indent=2)
        yaml_text = yaml.dump(json_data, allow_unicode=True, default_flow_style=False)

        last_row = ws.max_row
        output_row = last_row + offset
        title_row = output_row - 1

        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        title_fill = PatternFill(fill_type='solid', start_color='E6E6FA')

        self._set_title_cell(ws, title_row, 1, "出力JSON", title_fill, thin_border)
        self._set_title_cell(ws, title_row, 2, "出力YAML", title_fill, thin_border)

        self._setup_output_cell(ws, output_row, 1, json_text, thin_border)
        self._setup_output_cell(ws, output_row, 2, yaml_text, thin_border)

    def save(self, output_file):
        """Workbook を指定のファイル名で保存します。"""
        self.wb.save(output_file)
        print(f"新しい出力Excelファイルを作成しました： {output_file}")

    def run(self, output_file):
        """
        全体の処理を実行します。
          1. 対象シートの複製
          2. データマーカーの変換
          3. JSONデータへの変換
          4. JSON/YAMLの出力
          5. ファイルの保存
        """
        self.duplicate_sheet()
        self.transform_data_markers()
        json_data = self.convert_sheet_to_json()
        if json_data is None:
            print("JSON変換に失敗しました。")
            sys.exit(1)
        self.output_json_yaml(json_data)
        self.save(output_file)

def print_usage():
    print("使い方: python script.py 入力ファイル.xlsx [出力ファイル.xlsx] [-r|--replace]")
    print("  ・出力ファイル名を省略した場合、入力ファイル名に _出力日時 を付加した名前になります。")
    print("  ・-r または --replace を指定すると、処理後に出力ファイルで入力ファイルを差し替えます。")

if __name__ == "__main__":
    if any(arg in ("-h", "--help") for arg in sys.argv):
        print_usage()
        sys.exit(0)

    replace_flag = False
    positional_args = []
    for arg in sys.argv[1:]:
        if arg in ("-r", "--replace"):
            replace_flag = True
        else:
            positional_args.append(arg)

    if len(positional_args) < 1:
        print_usage()
        sys.exit(1)

    input_filename = positional_args[0]
    if len(positional_args) >= 2:
        output_filename = positional_args[1]
    else:
        base, ext = os.path.splitext(input_filename)
        dt_str = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_filename = f"{base}_{dt_str}.xlsx"

    converter = ExcelJsonConverter(input_filename)
    converter.run(output_filename)

    if replace_flag:
        try:
            os.replace(output_filename, input_filename)
            print(f"入力ファイル {input_filename} を更新しました。")
        except Exception as e:
            print(f"入力ファイルの差し替えに失敗しました: {e}")
