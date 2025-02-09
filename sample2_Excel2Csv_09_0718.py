import openpyxl
import csv
import sys
from openpyxl.utils import get_column_letter, column_index_from_string
from typing import Optional, Tuple, List, NamedTuple

class TargetRange(NamedTuple):
    start_row: int
    end_row: int
    target_cols: List[int]

def process_excel(
    excel_file: str,
    sheet_name: Optional[str] = None,
    start_cell: str = 'A1'
) -> None:
    """
    エクセルファイルを処理してCSVに変換する
    
    Args:
        excel_file (str): 入力エクセルファイル名
        sheet_name (Optional[str]): 処理対象のシート名（デフォルトはNone）
        start_cell (str): 解析開始セル（デフォルトは'A1'）
    """
    try:
        # エクセルファイルを開く
        workbook = openpyxl.load_workbook(excel_file, data_only=True)
        
        # シート名の決定
        if sheet_name is None:
            if len(workbook.sheetnames) == 1:
                sheet_name = workbook.sheetnames[0]
            else:
                sheet_name = 'LAYOUT'
        
        if sheet_name not in workbook.sheetnames:
            raise ValueError(f"Sheet '{sheet_name}' not found in workbook")
        
        sheet = workbook[sheet_name]
        
        # 開始セルの列と行を取得
        start_col = column_index_from_string(start_cell[0])
        layout_start_row = int(start_cell[1:])

        # LAYOUTの行を探す
        layout_row = find_layout_row(sheet, layout_start_row)
        if layout_row is None:
            raise ValueError("LAYOUT row not found")
        layout_last_row = find_layout_last_row(sheet, layout_start_row)

        # 対象となる列を特定（1列目は常に含める）
        target_cols = [start_col] + find_target_columns(sheet, start_col, layout_row)
        
        csv_layout = create_csv_layout(sheet, layout_start_row, target_cols)

        # データ範囲を特定
        target_range = find_target_range(sheet, layout_last_row, target_cols)
        
        # CSVデータの作成
        csv_data = create_csv_data(sheet, target_range)
        
        # CSVファイルの出力
        output_filename = f"{excel_file.rsplit('.', 1)[0]}.csv"
        write_csv(output_filename,csv_layout + csv_data)
        
    except Exception as e:
        print(f"Error processing Excel file: {str(e)}")
        sys.exit(1)

def find_layout_row(sheet, start_row: int) -> Optional[int]:
    """
    LAYOUT行を探す
    """
    max_row = sheet.max_row
    for row in range(start_row, max_row + 1):
        cell_value = str(sheet.cell(row=row, column=1).value or '').upper()
        if cell_value == 'LAYOUT':
            return row
    return None

def find_layout_last_row(sheet, start_row: int) -> Optional[int]:
    """
    LAYOUT行の最後の行を探す
    """
    max_row = sheet.max_row
    layout_rows = []
    for row in range(start_row, max_row + 1):
        cell_value = str(sheet.cell(row=row, column=1).value or '').upper()
        if cell_value == 'LAYOUT':
            layout_rows.append(row)
    return layout_rows[-1]

def create_csv_layout(sheet, layout_start_row: int, target_cols: List[int]) -> List[str]:
    """
    LAYOUT行の内容をCSV形式に変換する
    """
    max_row = sheet.max_row
    layout_rows = []
    for row in range(layout_start_row, max_row + 1):
        row_data = []
        cell_value = str(sheet.cell(row=row, column=1).value or '').upper()
        if cell_value == 'LAYOUT':
            for col in target_cols:
                if col == target_cols[0]:  # 1列目は特別処理
                    value = process_layout_cell_value(sheet, row, col, check_prefix=False)
                else:
                    value = process_layout_cell_value(sheet, row, col, check_prefix=True)
                #print(f'row:{row} col:{col} value:{value}')

                row_data.append(value)
            layout_rows.append(row_data)
    return layout_rows



def find_target_columns(sheet, start_col: int, layout_row: int) -> List[int]:
    """
    プレフィックスが#である値を含む列を特定する

    """
    target_cols = []
    max_col = sheet.max_column
    
    for col in range(start_col + 1, max_col + 1):  # 2列目以降を検査
        cell = sheet.cell(row=layout_row, column=col)
        
        # 結合セルの処理
        if cell.coordinate in sheet.merged_cells:
            merge_range = next(range for range in sheet.merged_cells.ranges if cell.coordinate in range)
            value = sheet.cell(row=merge_range.min_row, column=merge_range.min_col).value
        else:
            value = cell.value
            
        if value and isinstance(value, str) and value.startswith('#'):
            target_cols.append(col)
    
    return target_cols

def find_target_range(sheet, layout_last_row: int, target_cols: List[int]) -> TargetRange:
    """
    対象となるデータの範囲を特定する
    """
    max_row = sheet.max_row
    start_keywords = ['START']
    skip_keywords = ['NONE', 'NOT', 'NO']
    end_keywords = ['END', 'FINISH', 'FIN']
    
    # 開始行の特定
    start_row = layout_last_row + 1  # デフォルトはLAYOUTの次の行
    for row in range(layout_last_row + 1, max_row + 1):
        cell_value = str(sheet.cell(row=row, column=1).value or '').upper()
        if cell_value in start_keywords:
            start_row = row

            break
    
    # 終了行の特定
    end_row = max_row
    for row in range(start_row, max_row + 1):
        cell_value = str(sheet.cell(row=row, column=1).value or '').upper()
        if cell_value in end_keywords:
            end_row = row - 1
            break
    
    return TargetRange(start_row, end_row, target_cols)

def create_csv_data(sheet, target_range: TargetRange) -> List[List[str]]:
    """
    CSVデータを作成する
    """
    csv_data = []
    skip_keywords = ['NONE', 'NOT', 'NO']
    
    for row in range(target_range.start_row, target_range.end_row + 1):
        # スキップ条件のチェック
        first_col_value = str(sheet.cell(row=row, column=1).value or '').lower()
        if first_col_value in skip_keywords:
            continue
            
        row_data = []
        for col in target_range.target_cols:
            #cell = sheet.cell(row=row, column=col)
            if col == target_range.target_cols[0]:  # 1列目は特別処理
                value = process_cell_value(sheet, row, col, check_prefix=False)
            else:
                value = process_cell_value(sheet, row, col, check_prefix=True)
            row_data.append(value)
        csv_data.append(row_data)
    
    return csv_data

def process_layout_cell_value(sheet, row: int, col: int, check_prefix: bool = True) -> str:
    """
    セルの値を処理する
    """
    cell = sheet.cell(row=row, column=col)
    
    # 連結セルのチェック
    if cell.coordinate in sheet.merged_cells:
        merge_range = next(range for range in sheet.merged_cells.ranges if cell.coordinate in range)
        if cell.coordinate == merge_range.start_cell.coordinate:
            value = cell.value
            if value.startswith('#'):
                value = value[1:]  # #を除去
            else:
                value = "" # #で始まるセルでなければ除外する
        else:
            return '<'
    
    # 通常のセル
    value = cell.value
    if check_prefix and value and isinstance(value, str):
        if value.startswith('#'):
            value = value[1:]  # #を除去
        else:
            value = "" # #で始まるセルでなければ除外する
    return str(value) if value is not None else ''


def process_cell_value(sheet, row: int, col: int, check_prefix: bool = True) -> str:
    """
    セルの値を処理する
    """
    cell = sheet.cell(row=row, column=col)
    
    # 連結セルのチェック
    if cell.coordinate in sheet.merged_cells:
        merge_range = next(range for range in sheet.merged_cells.ranges if cell.coordinate in range)
        if cell.coordinate == merge_range.start_cell.coordinate:
            value = cell.value
            if check_prefix and value and isinstance(value, str) and value.startswith('#'):
                value = value[1:]  # #を除去
            return value
        else:
            return '<'
    
    # 通常のセル
    value = cell.value
    if check_prefix and value and isinstance(value, str) and value.startswith('#'):
        value = value[1:]  # #を除去
    return str(value) if value is not None else ''

def write_csv(filename: str, data: List[List[str]]) -> None:
    """
    データをCSVファイルに書き出す
    """
    with open(filename, 'w', newline='', encoding='utf-8') as f:
        writer = csv.writer(f)
        writer.writerows(data)

if __name__ == "__main__":
    import argparse
    
    parser = argparse.ArgumentParser(description='Convert Excel file to CSV with specific rules')
    parser.add_argument('excel_file', help='Input Excel file name')
    parser.add_argument('--sheet', help='Target sheet name (default: LAYOUT if multiple sheets exist)')
    parser.add_argument('--start', default='A1', help='Start cell (default: A1)')
    
    args = parser.parse_args()
    
    process_excel(args.excel_file, args.sheet, args.start)