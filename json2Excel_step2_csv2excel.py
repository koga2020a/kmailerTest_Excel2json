import argparse
from openpyxl import Workbook
from openpyxl.styles import Alignment, PatternFill

def main():
    parser = argparse.ArgumentParser(
        description="TXT ファイルを読み込み、Excel に変換します。"
    )
    parser.add_argument("input_txt", help="入力 TXT ファイル")
    parser.add_argument("output_xlsx", nargs="?", help="出力 Excel ファイル（省略時は入力ファイル名の拡張子を .xlsx にしたもの）")
    parser.add_argument("--separate", action="store_true",
                        help="LAYOUT 行でのセル結合を行わず、そのまま出力する")
    parser.add_argument("--no-color", action="store_true",
                        help="結合セルの背景色付けを行わない")
    args = parser.parse_args()

    input_txt = args.input_txt
    output_xlsx = args.output_xlsx if args.output_xlsx else input_txt.rsplit('.', 1)[0] + '_output.xlsx'
    separate = args.separate
    no_color = args.no_color

    # TXT ファイルを読み込み、各行をタブ区切りで分割する
    with open(input_txt, 'r', encoding='utf-8') as f:
        lines = f.read().splitlines()
    data = [line.split("\t") for line in lines]

    wb = Workbook()
    ws = wb.active

    # TXT の内容をセルに書き込む（行・列とも 1-indexed）
    for row_index, row in enumerate(data, start=1):
        for col_index, cell_value in enumerate(row, start=1):
            ws.cell(row=row_index, column=col_index, value=cell_value)

    # LAYOUT 行の処理
    # ※レイアウト行は先頭セルが "LAYOUT" である行とします。
    #  ここでは、LAYOUT 行の2列目以降のセルを中央揃えにし、
    #  さらに ( --separate 指定がない場合 ) 連続するセルグループを結合し、
    #  結合セルには背景色を設定します（--no-color 指定で色付けを抑制）。
    candidate_colors = ["FFCCCC", "CCFFCC", "CCCCFF", "FFFFCC", "FFCCFF", "CCFFFF"]
    # ※候補は薄いパステル調の色です。

    for row_index, row in enumerate(data, start=1):
        if row and row[0] == "LAYOUT":
            # LAYOUT 行の2列目以降（Excel では列2～）を中央揃えにする
            for col_index in range(2, len(row) + 1):
                cell = ws.cell(row=row_index, column=col_index)
                cell.alignment = Alignment(horizontal="center")

            if not separate:
                # 結合処理（結合対象は、先頭2セル("LAYOUT" とその次のセル)を除いた、リスト上の index 2 以降）
                col = 2
                last_merged_color = None  # 同じ行内で直前に結合したセルの背景色
                while col < len(row):
                    cell_val = row[col]
                    # 結合グループの先頭は、値が空でなく、"<" でないセルとする
                    if cell_val and cell_val != "<":
                        group_start = col
                        group_end = col
                        # 直後のセルが "<" なら同じグループとみなす
                        while group_end + 1 < len(row) and row[group_end + 1] == "<":
                            group_end += 1
                        if group_end > group_start:
                            # Python のリスト index を Excel の列番号（1-indexed）に変換して merge_cells
                            ws.merge_cells(start_row=row_index,
                                           start_column=group_start + 1,
                                           end_row=row_index,
                                           end_column=group_end + 1)
                            # 背景色付け（--no-color オプションがなければ）
                            if not no_color:
                                # 前グループの色と異なる色を候補から選択する
                                chosen_color = None
                                for color in candidate_colors:
                                    if color != last_merged_color:
                                        chosen_color = color
                                        break
                                if chosen_color is None:
                                    chosen_color = candidate_colors[0]
                                last_merged_color = chosen_color
                                cell = ws.cell(row=row_index, column=group_start + 1)
                                cell.fill = PatternFill(start_color=chosen_color,
                                                        end_color=chosen_color,
                                                        fill_type="solid")
                        col = group_end + 1
                    else:
                        col += 1

    wb.save(output_xlsx)
    print(f"Excelファイルが作成されました: {output_xlsx}")

if __name__ == "__main__":
    main()
