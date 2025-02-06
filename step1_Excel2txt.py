#!/usr/bin/env python
# -*- coding: utf-8 -*-

import sys
from openpyxl import load_workbook

def get_target_columns(sheet):
    """
    シート内の HEAD 行から、3列目以降で先頭が '#' もしくは '[' のセルがある列（1-indexed）を収集。
    さらに1列目、2列目は必ず対象とする。
    """
    target_cols = set()
    for row in sheet.iter_rows(values_only=True):
        if not row:
            continue
        if row[0] == "HEAD":
            # 3列目以降 (インデックスは2以降) をチェック
            for i, cell in enumerate(row[2:], start=3):
                if cell is not None:
                    s = str(cell)
                    if s.startswith("#") or s.startswith("["):
                        target_cols.add(i)
    # 1,2列目は必ず含む
    target_cols.add(1)
    target_cols.add(2)
    return sorted(target_cols)

def format_head_row(row, target_cols):
    """
    HEAD行の場合、3列目以降で先頭が '#' の場合は除去して出力用のリストを作る。
    """
    out_fields = []
    for col in target_cols:
        # 列番号は1-indexed、Pythonではcol-1
        val = row[col-1] if col-1 < len(row) else ""
        if col >= 3 and val is not None:
            s = str(val)
            if s.startswith("#"):
                s = s[1:]
            out_fields.append(s)
        else:
            out_fields.append("" if val is None else str(val))
    return out_fields

def format_data_row(row, target_cols, force_first=False):
    """
    DATA系の行の場合、force_first が True のときは1列目を "DATA" に強制する
    """
    out_fields = []
    for col in target_cols:
        if force_first and col == 1:
            out_fields.append("DATA")
        else:
            val = row[col-1] if col-1 < len(row) else ""
            out_fields.append("" if val is None else str(val))
    return out_fields

def main():
    if len(sys.argv) < 2:
        print("Usage: {} excel_file [sheet_name]".format(sys.argv[0]))
        sys.exit(1)

    excel_file = sys.argv[1]
    specified_sheet = sys.argv[2] if len(sys.argv) > 2 else None

    # Excelファイルを読み込み（計算結果など data_only=True で取得）
    try:
        wb = load_workbook(excel_file, data_only=True)
    except Exception as e:
        sys.exit("Excelファイルの読み込みに失敗しました: {}".format(e))

    # 対象シートの選択
    sheet = None
    if specified_sheet:
        if specified_sheet in wb.sheetnames:
            sheet = wb[specified_sheet]
        else:
            sys.exit("指定されたシート '{}' が見つかりません。".format(specified_sheet))
    else:
        if len(wb.sheetnames) == 1:
            sheet = wb.active
        else:
            if "HEAD" in wb.sheetnames:
                sheet = wb["HEAD"]
            else:
                sys.exit("複数シートありますが、デフォルトの 'HEAD' シートが見つかりません。")

    # 出力対象の列をHEAD行から決定
    target_cols = get_target_columns(sheet)
    # ※ 出力する列番号（1-indexed）の一覧例: [1, 2, 5, 7] など

    output_file = excel_file.rsplit(".", 1)[0] + ".txt"
    try:
        fout = open(output_file, "w", encoding="utf-8")
    except Exception as e:
        sys.exit("出力ファイルのオープンに失敗しました: {}".format(e))

    in_data_block = False  # DATA_START ～ DATA_END ブロック内かどうか

    # シートの全行を処理
    for row in sheet.iter_rows(values_only=True):
        if not row:
            continue
        cell_a = row[0]  # 1列目の値
        if cell_a is None:
            continue

        if cell_a == "HEAD":
            # HEAD行は対象。出力時は3列目以降で先頭 '#' があれば除去
            out_fields = format_head_row(row, target_cols)
            fout.write("\t".join(out_fields) + "\n")
        elif cell_a == "DATA":
            # DATA行はそのまま出力
            out_fields = format_data_row(row, target_cols)
            fout.write("\t".join(out_fields) + "\n")
        elif cell_a == "DATA_START":
            # DATA_START行～DATA_END行または最終行までをひとまとめに対象とする
            in_data_block = True
            # DATA_START行も出力。1列目は強制的に "DATA" にする
            out_fields = format_data_row(row, target_cols, force_first=True)
            fout.write("\t".join(out_fields) + "\n")
        elif cell_a == "DATA_END":
            # DATA_END はブロックの終了。出力せずフラグを下ろす
            in_data_block = False
        elif cell_a == "NONE":
            # NONE の行は出力対象外
            continue
        else:
            # もし DATA_START ブロック内であれば、DATA_ENDに達する前の行が対象
            if in_data_block:
                out_fields = format_data_row(row, target_cols, force_first=True)
                fout.write("\t".join(out_fields) + "\n")
            # それ以外は対象外
    fout.close()
    print("出力完了: {}".format(output_file))

if __name__ == '__main__':
    main()
