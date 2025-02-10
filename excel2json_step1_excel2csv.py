import openpyxl
import csv
import sys
from openpyxl.utils import get_column_letter, column_index_from_string
from typing import Optional, List, NamedTuple

# 対象データ範囲を表す
class TargetRange(NamedTuple):
    start_row: int
    end_row: int
    target_cols: List[int]

# Excelファイルおよびシートの読み込みと開始位置の情報を保持する
class ExcelDocument:
    def __init__(self, excel_file: str, sheet_name: Optional[str] = None, start_cell: str = 'A1'):
        """
        Excelファイルの読み込みと対象シート、解析開始セルの設定を行う。
        """
        self.excel_file = excel_file
        self.sheet_name = sheet_name
        self.start_cell = start_cell
        self.workbook = openpyxl.load_workbook(excel_file, data_only=True)
        self._select_sheet()
        self.start_col = column_index_from_string(self.start_cell[0])
        self.start_row = int(self.start_cell[1:])

    def _select_sheet(self) -> None:
        # シート名の決定
        if self.sheet_name is None:
            if len(self.workbook.sheetnames) == 1:
                self.sheet_name = self.workbook.sheetnames[0]
            else:
                self.sheet_name = 'LAYOUT'
        if self.sheet_name not in self.workbook.sheetnames:
            raise ValueError(f"Sheet '{self.sheet_name}' not found in workbook")
        self.sheet = self.workbook[self.sheet_name]

# セルの値の処理を担当するヘルパークラス
class CellProcessor:
    @staticmethod
    def process_layout_cell(sheet, row: int, col: int, check_prefix: bool = True) -> str:
        """
        LAYOUT行内のセルの値を処理する。結合セルの場合は左上セルの値を参照し、
        値が文字列の場合は'#'プレフィックスの除去を行う。
        """
        cell = sheet.cell(row=row, column=col)
        if cell.coordinate in sheet.merged_cells:
            merge_range = next(r for r in sheet.merged_cells.ranges if cell.coordinate in r)
            if cell.coordinate == merge_range.start_cell.coordinate:
                value = cell.value
                if check_prefix and value and isinstance(value, str):
                    if value.startswith('#'):
                        value = value[1:]
                    else:
                        value = ""
            else:
                start_cell_value = sheet.cell(row=merge_range.min_row, column=merge_range.min_col).value
                if check_prefix and start_cell_value and isinstance(start_cell_value, str) and start_cell_value.startswith('#'):
                    return '<'
                return ''
        else:
            value = cell.value
            if check_prefix and value and isinstance(value, str):
                if value.startswith('#'):
                    value = value[1:]
                else:
                    value = ""
        return str(value) if value is not None else ''

    @staticmethod
    def process_data_cell(sheet, row: int, col: int, check_prefix: bool = True) -> str:
        """
        データ部門のセルの値を処理する。結合セルの縦横の状況に応じて、
        先頭セル以外は空文字を返すなどの処理を行う。
        """
        cell = sheet.cell(row=row, column=col)
        if cell.coordinate in sheet.merged_cells:
            merge_range = next(r for r in sheet.merged_cells.ranges if cell.coordinate in r)
            is_vertical_merge = merge_range.min_col == merge_range.max_col
            if is_vertical_merge:
                if cell.coordinate == merge_range.start_cell.coordinate:
                    value = cell.value
                else:
                    return ''
            else:
                if row == merge_range.min_row:
                    value = sheet.cell(row=merge_range.min_row, column=merge_range.min_col).value
                else:
                    return ''
        else:
            value = cell.value

        if check_prefix and value and isinstance(value, str) and value.startswith('#'):
            value = value[1:]
        return str(value) if value is not None else ''

# LAYOUT部門の処理を担当するクラス
class LayoutSection:
    def __init__(self, sheet, start_row: int, start_col: int):
        """
        LAYOUT部門の開始行・列を元に、LAYOUT行と対象列の情報を抽出する。
        """
        self.sheet = sheet
        self.start_row = start_row
        self.start_col = start_col
        self.layout_rows = self._find_layout_rows()
        if not self.layout_rows:
            raise ValueError("LAYOUT行が見つかりません")
        self.last_layout_row = self.layout_rows[-1]
        self.target_cols = self._find_target_columns()
        # 開始セルの列は常に先頭に含める
        if self.start_col not in self.target_cols:
            self.target_cols.insert(0, self.start_col)
        else:
            self.target_cols.sort(key=lambda x: (0 if x == self.start_col else x))

    def _find_layout_rows(self) -> List[int]:
        """
        開始セルから最終行までの範囲で、1列目が'LAYOUT'と一致する行番号を返す。
        """
        layout_rows = []
        max_row = self.sheet.max_row
        for row in range(self.start_row, max_row + 1):
            value = str(self.sheet.cell(row=row, column=1).value or '').upper()
            if value == 'LAYOUT':
                layout_rows.append(row)
        return layout_rows

    def _find_target_columns(self) -> List[int]:
        """
        LAYOUT行の各セルから、値が'#'で始まる列番号を特定する。
        """
        target_set = set()
        max_col = self.sheet.max_column
        for layout_row in self.layout_rows:
            for col in range(self.start_col + 1, max_col + 1):
                cell = self.sheet.cell(row=layout_row, column=col)
                if cell.coordinate in self.sheet.merged_cells:
                    merge_range = next(r for r in self.sheet.merged_cells.ranges if cell.coordinate in r)
                    value = self.sheet.cell(row=merge_range.min_row, column=merge_range.min_col).value
                else:
                    value = cell.value
                if value and isinstance(value, str) and value.startswith('#'):
                    target_set.add(col)
        return sorted(target_set)

    def create_csv_layout(self) -> List[List[str]]:
        """
        LAYOUT行の内容を抽出し、CSV出力用の2次元リストを作成する。
        """
        csv_layout = []
        max_row = self.sheet.max_row
        for row in range(self.start_row, max_row + 1):
            if str(self.sheet.cell(row=row, column=1).value or '').upper() == 'LAYOUT':
                row_data = []
                for col in self.target_cols:
                    if col == self.target_cols[0]:
                        value = CellProcessor.process_layout_cell(self.sheet, row, col, check_prefix=False)
                    else:
                        value = CellProcessor.process_layout_cell(self.sheet, row, col, check_prefix=True)
                    row_data.append(value)
                # 1列目以外が全て空の場合は無視する
                if all(x == '' for x in row_data[1:]):
                    continue
                csv_layout.append(row_data)
        return csv_layout

# データ部門の処理を担当するクラス
class DataSection:
    def __init__(self, sheet, last_layout_row: int, target_cols: List[int]):
        """
        LAYOUT部門の最終行と対象列の情報をもとに、データ部門の対象範囲を決定する。
        """
        self.sheet = sheet
        self.last_layout_row = last_layout_row
        self.target_cols = target_cols
        self.target_range = self._find_target_range()

    def _find_target_range(self) -> TargetRange:
        """
        データ部門の開始行（'START'の有無により決定）と終了行（'END'などのキーワードの直前）を特定する。
        """
        max_row = self.sheet.max_row
        start_keywords = ['START']
        end_keywords = ['END', 'FINISH', 'FIN']
        start_row = self.last_layout_row + 1
        for row in range(self.last_layout_row + 1, max_row + 1):
            if str(self.sheet.cell(row=row, column=1).value or '').upper() in start_keywords:
                start_row = row
                break
        end_row = max_row
        for row in range(start_row, max_row + 1):
            if str(self.sheet.cell(row=row, column=1).value or '').upper() in end_keywords:
                end_row = row - 1
                break
        return TargetRange(start_row, end_row, self.target_cols)

    def create_csv_data(self) -> List[List[str]]:
        """
        データ部門の対象範囲から、CSV出力用の2次元リストを作成する。
        """
        csv_data = []
        skip_keywords = ['none', 'not', 'no']
        for row in range(self.target_range.start_row, self.target_range.end_row + 1):
            if str(self.sheet.cell(row=row, column=1).value or '').lower() in skip_keywords:
                continue
            row_data = []
            for col in self.target_range.target_cols:
                if col == self.target_range.target_cols[0]:
                    value = CellProcessor.process_data_cell(self.sheet, row, col, check_prefix=False)
                else:
                    value = CellProcessor.process_data_cell(self.sheet, row, col, check_prefix=True)
                row_data.append(value)
            csv_data.append(row_data)
        return csv_data

# CSVファイル出力を担当するクラス
class CSVWriter:
    @staticmethod
    def write_csv(filename: str, data: List[List[str]]) -> None:
        """
        与えられた2次元リストのデータをCSVファイルに出力する。
        """
        with open(filename, 'w', newline='', encoding='utf-8') as f:
            writer = csv.writer(f)
            writer.writerows(data)

# 全体の処理を統括するコンバータークラス
class ExcelToCSVConverter:
    def __init__(self, excel_file: str, sheet_name: Optional[str] = None, start_cell: str = 'A1'):
        self.document = ExcelDocument(excel_file, sheet_name, start_cell)

    def convert(self) -> None:
        """
        ExcelファイルのLAYOUT部門およびデータ部門を処理し、CSVファイルとして出力する。
        """
        # LAYOUT部門の処理
        layout_section = LayoutSection(self.document.sheet, self.document.start_row, self.document.start_col)
        csv_layout = layout_section.create_csv_layout()

        # データ部門の処理
        data_section = DataSection(self.document.sheet, layout_section.last_layout_row, layout_section.target_cols)
        csv_data = data_section.create_csv_data()

        # CSVファイルの出力
        output_filename = f"{self.document.excel_file.rsplit('.', 1)[0]}.csv"
        CSVWriter.write_csv(output_filename, csv_layout + csv_data)
        print(f"CSVファイル '{output_filename}' を作成しました。")

if __name__ == "__main__":
    import argparse

    parser = argparse.ArgumentParser(
        description='Excelファイルを特定ルールに従いCSVに変換する'
    )
    parser.add_argument('excel_file', help='入力Excelファイル名')
    parser.add_argument('--sheet', help='対象シート名（デフォルト：LAYOUTまたは唯一のシート）')
    parser.add_argument('--start', default='A1', help="解析開始セル（例：A1）")
    
    args = parser.parse_args()

    try:
        converter = ExcelToCSVConverter(args.excel_file, args.sheet, args.start)
        converter.convert()
    except Exception as e:
        print(f"Error processing Excel file: {str(e)}")
        sys.exit(1)
